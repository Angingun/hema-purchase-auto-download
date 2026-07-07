import hashlib
import logging
import time
import os
from datetime import datetime, timedelta

logger = logging.getLogger(__name__)


def setup_logging(log_dir: str):
    """初始化日志（同时输出到控制台和文件）"""
    os.makedirs(log_dir, exist_ok=True)
    log_file = os.path.join(log_dir, f"run_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log")
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s [%(levelname)s] %(message)s",
        handlers=[
            logging.StreamHandler(),
            logging.FileHandler(log_file, encoding="utf-8"),
        ],
    )
    logger.info(f"日志文件: {log_file}")


def get_date_range(end_date_str: str = None, days_back: int = 7):
    """
    返回 (start_date, end_date) 字符串，格式 YYYY-MM-DD
    end_date 默认今天，start_date = end_date - days_back 天
    """
    if end_date_str:
        end = datetime.strptime(end_date_str, "%Y-%m-%d")
    else:
        end = datetime.today()
    start = end - timedelta(days=days_back)
    return start.strftime("%Y-%m-%d"), end.strftime("%Y-%m-%d")


def wait_for_new_file(download_dir: str, before_files: set, timeout: int = 60) -> str | None:
    """等待下载目录出现新文件，返回文件路径"""
    deadline = time.time() + timeout
    while time.time() < deadline:
        current = set(os.listdir(download_dir))
        new_files = current - before_files
        # 过滤掉 .crdownload 临时文件
        completed = [f for f in new_files if not f.endswith(".crdownload")]
        if completed:
            path = os.path.join(download_dir, completed[0])
            logger.info(f"下载完成: {path}")
            return path
        time.sleep(1)
    logger.warning("等待下载超时")
    return None


DownloadResult = dict[str, object]
"""字段约定：
    ok: bool           — 下载是否成功
    path: str | None   — 文件完整路径
    filename: str | None — 文件名
    size_bytes: int | None — 最终文件大小
    stable: bool       — 文件大小是否已稳定
    error: str | None  — 失败原因
"""


def wait_download_complete(download_dir: str, before_files: set[str],
                           timeout: int = 120) -> DownloadResult:
    """等待下载完成，包含文件大小稳定性校验。

    规则：
      1. 忽略 .crdownload 临时文件
      2. 等待新文件出现
      3. 等待文件大小连续两次采样相同（稳定）
      4. 文件大小为 0 视为失败

    返回 DownloadResult。
    """
    deadline = time.time() + timeout
    last_size: int | None = None
    stable_count = 0
    new_file_path: str | None = None
    new_filename: str | None = None

    while time.time() < deadline:
        current = set(os.listdir(download_dir))
        new_files = current - before_files
        completed = sorted(f for f in new_files
                          if not f.endswith('.crdownload'))

        if completed:
            new_filename = completed[0]
            new_file_path = os.path.join(download_dir, new_filename)

            try:
                size = os.path.getsize(new_file_path)
            except OSError:
                size = 0

            if size == 0:
                logger.warning("Downloaded file is empty: %s", new_filename)
                time.sleep(1)
                continue

            if last_size is not None and last_size == size:
                stable_count += 1
                if stable_count >= 2:
                    logger.info("Download stable: %s (%d bytes)",
                                new_filename, size)
                    return {
                        'ok': True,
                        'path': new_file_path,
                        'filename': new_filename,
                        'size_bytes': size,
                        'stable': True,
                        'error': None,
                    }
            else:
                stable_count = 0
            last_size = size

        remaining = deadline - time.time()
        if remaining > 0:
            time.sleep(min(1.0, remaining))

    # 超时
    if new_file_path and os.path.exists(new_file_path):
        size = os.path.getsize(new_file_path)
        return {
            'ok': False,
            'path': new_file_path,
            'filename': new_filename,
            'size_bytes': size,
            'stable': False,
            'error': f'Download did not stabilize within {timeout}s',
        }
    return {
        'ok': False,
        'path': None,
        'filename': None,
        'size_bytes': None,
        'stable': False,
        'error': f'No new file appeared within {timeout}s',
    }


ExcelVerifyResult = dict[str, object]
"""字段约定：
    ok: bool               — 可解析且通过基本校验
    path: str              — 被校验文件路径
    sheet_names: list[str] — 工作表名称列表
    row_count: int | None  — 第一个有数据的 sheet 的行数
    column_count: int | None — 第一个有数据的 sheet 的列数
    error: str | None      — 失败原因
"""


def verify_excel_file(path: str) -> ExcelVerifyResult:
    """校验 Excel 文件可解析并返回文件元数据。

    要求：
      - 文件必须存在
      - 必须能用 openpyxl 解析
      - 至少包含 1 个 sheet
      - 返回第一个非空 sheet 的行数和列数

    依赖：openpyxl（已在 requirements.txt 中添加）
    """
    result: ExcelVerifyResult = {
        'ok': False,
        'path': path,
        'sheet_names': [],
        'row_count': None,
        'column_count': None,
        'order_ids': [],
        'order_hash': None,
        'error': None,
    }

    if not os.path.exists(path):
        result['error'] = f'File not found: {path}'
        return result

    try:
        import openpyxl
    except ImportError:
        result['error'] = 'openpyxl is not installed (run: pip install openpyxl)'
        return result

    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        result['sheet_names'] = wb.sheetnames

        if not wb.sheetnames:
            result['error'] = 'Excel file has no sheets'
            wb.close()
            return result

        # 读取第一个 sheet 的行列数，并抽取采购单号集合
        ws = wb[wb.sheetnames[0]]
        rows = 0
        cols = 0
        order_ids = []
        seen = set()
        for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
            values = ['' if v is None else str(v).strip() for v in row]
            rows += 1
            if cols == 0:
                cols = len(values)
            if row_idx == 0:
                continue
            order_id = next((v for v in values if v.startswith('HPO')), '')
            if order_id and order_id not in seen:
                seen.add(order_id)
                order_ids.append(order_id)
        result['row_count'] = rows
        result['column_count'] = cols
        result['order_ids'] = order_ids
        result['order_hash'] = hashlib.sha256(
            '\n'.join(sorted(set(order_ids))).encode('utf-8')
        ).hexdigest() if order_ids else None
        result['ok'] = True
        wb.close()

    except Exception as e:
        result['error'] = f'Failed to parse Excel: {e}'

    return result


